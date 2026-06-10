#!/usr/bin/env python3
"""Convert a World Cup pool Excel workbook into normalized player JSON."""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


PROJECT_ROOT = Path(__file__).resolve().parent.parent
SHEET_NAME = "Pool"
PHASES = (
    "group_stage",
    "round_of_32",
    "round_of_16",
    "quarter_finals",
    "semi_finals",
    "third_place",
    "final",
)
EXPECTED_MATCH_COUNTS = {
    "group_stage": 72,
    "round_of_32": 16,
    "round_of_16": 8,
    "quarter_finals": 4,
    "semi_finals": 2,
    "third_place": 1,
    "final": 1,
}
EXPECTED_QUALIFIED_COUNTS = {
    "round_of_32": 32,
    "round_of_16": 16,
    "quarter_finals": 8,
    "semi_finals": 4,
    "third_place": 2,
    "final": 2,
}

PHASE_HEADERS = {
    "FASE DE GRUPOS": "group_stage",
    "ENFRENTAMIENTOS DIECISEISAVOS": "round_of_32",
    "ENFRENTAMIENTOS OCTAVOS": "round_of_16",
    "ENFRENTAMIENTOS CUARTOS": "quarter_finals",
    "ENFRENTAMIENTOS SEMIFINALES": "semi_finals",
    "3O 4O PUESTO": "third_place",
    "ENFRENTAMIENTO FINAL": "final",
}

HONOR_LABELS = {
    "CAMPEON": "champion",
    "SUBCAMPEON": "runner_up",
    "3O PUESTO": "third_place",
}

AWARD_LABELS = {
    "BOTA DE ORO": "golden_boot",
    "BOTA DE PLATA": "silver_boot",
    "BOTA DE BRONCE": "bronze_boot",
    "BALON DE ORO": "golden_ball",
    "BALON DE PLATA": "silver_ball",
    "BALON DE BRONCE": "bronze_ball",
}

QUALIFIED_LABELS = {
    "DIECISEISAVOFINALISTA": "round_of_32",
    "OCTAVOFINALISTA": "round_of_16",
    "CUARTOFINALISTA": "quarter_finals",
    "SEMIFINALISTA": "semi_finals",
    "3O Y 4O PUESTO": "third_place",
    "FINALISTA": "final",
}

SCORE_RE = re.compile(
    r"(?P<outcome>[12Xx])\s*\|\s*(?P<home>\d+)\s*-\s*(?P<away>\d+)(?P<tail>.*)$"
)
PENALTY_RES = (
    re.compile(
        r"^[\s·|,;/\-]*(?:PEN(?:ALES)?|P\.?|PKS?)\s*[:=]?\s*"
        r"(?P<home>\d+)\s*-\s*(?P<away>\d+)\s*$",
        re.IGNORECASE,
    ),
    re.compile(
        r"^[\s·|,;/\-]*\(\s*(?P<home>\d+)\s*-\s*(?P<away>\d+)"
        r"\s*(?:PEN(?:ALES)?|P\.?|PKS?)\s*\)\s*$",
        re.IGNORECASE,
    ),
    re.compile(
        r"^[\s·|,;/\-]*\(\s*(?:PEN(?:ALES)?|P\.?|PKS?)\s*"
        r"(?P<home>\d+)\s*-\s*(?P<away>\d+)\s*\)\s*$",
        re.IGNORECASE,
    ),
    re.compile(
        r"^[\s·|,;/\-]+\s*(?P<home>\d+)\s*-\s*(?P<away>\d+)\s*$",
        re.IGNORECASE,
    ),
    re.compile(
        r"^[\s·|,;/\-]*\(\s*(?P<home>\d+)\s*-\s*(?P<away>\d+)\s*\)\s*$",
        re.IGNORECASE,
    ),
)


@dataclass
class ConversionResult:
    player_id: str
    player_name: str
    output_path: Path
    match_count: int
    award_count: int
    warnings: list[str] = field(default_factory=list)


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalized_label(value: Any) -> str:
    value_text = unicodedata.normalize("NFKD", text(value))
    ascii_text = value_text.encode("ascii", "ignore").decode("ascii")
    cleaned = re.sub(r"[^A-Za-z0-9]+", " ", ascii_text).strip().upper()
    return re.sub(r"\b([134])(?:O|A)\b", r"\1O", cleaned)


def slugify(value: str) -> str:
    value_text = unicodedata.normalize("NFKD", value)
    ascii_text = value_text.encode("ascii", "ignore").decode("ascii").lower()
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_text).strip("-")
    return slug or "jugador"


def detect_player_name(excel_path: Path) -> str:
    """Read the player's private name from Pool!C5 without generating output."""
    excel_path = excel_path.expanduser().resolve()
    if not excel_path.is_file():
        raise FileNotFoundError(f"No existe el archivo Excel: {excel_path}")

    try:
        workbook = load_workbook(excel_path, data_only=True, read_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise RuntimeError(f"No se pudo abrir el Excel: {exc}") from exc

    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise RuntimeError(f'El Excel no contiene la hoja requerida "{SHEET_NAME}"')
        player_name = text(workbook[SHEET_NAME]["C5"].value)
    finally:
        workbook.close()

    if not player_name:
        raise RuntimeError("No se pudo detectar el jugador en Pool!C5")
    return player_name


def split_match(match_name: str) -> tuple[str | None, str | None]:
    parts = [part.strip() for part in match_name.split("-", 1)]
    if len(parts) == 2 and all(parts):
        return parts[0], parts[1]
    return None, None


def integer_value(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def build_penalty_lookup(workbook: Any) -> dict[tuple[str, str, int, int], tuple[int, int]]:
    """Read knockout penalties stored next to scores in the WORLDCUP sheet."""
    if "WORLDCUP" not in workbook.sheetnames:
        return {}

    sheet = workbook["WORLDCUP"]
    lookup: dict[tuple[str, str, int, int], tuple[int, int]] = {}
    for values in sheet.iter_rows(
        min_col=27,
        max_col=32,
        values_only=True,
    ):
        home_team = text(values[0])
        home_penalties = integer_value(values[1])
        home_score = integer_value(values[2])
        away_score = integer_value(values[3])
        away_penalties = integer_value(values[4])
        away_team = text(values[5])
        if (
            home_team
            and away_team
            and home_penalties is not None
            and away_penalties is not None
            and home_score is not None
            and away_score is not None
        ):
            lookup[(home_team, away_team, home_score, away_score)] = (
                home_penalties,
                away_penalties,
            )
    return lookup


def add_penalties(
    item: dict[str, Any],
    home_penalties: int,
    away_penalties: int,
    row: int,
    warnings: list[str],
) -> None:
    item["home_penalties"] = home_penalties
    item["away_penalties"] = away_penalties
    item["penalties"] = {
        "home_score": home_penalties,
        "away_score": away_penalties,
    }

    score_match = SCORE_RE.search(item["prediction_raw"])
    raw_without_penalties = (
        item["prediction_raw"][: score_match.end("away")]
        if score_match
        else item["prediction_raw"]
    )
    item["prediction_raw"] = (
        f"{raw_without_penalties}|PEN:{home_penalties}-{away_penalties}"
    )

    if home_penalties > away_penalties:
        item["qualified_team"] = item.get("home_team")
    elif away_penalties > home_penalties:
        item["qualified_team"] = item.get("away_team")
    else:
        warnings.append(
            f"Fila {row}: penales empatados en eliminatoria "
            f"{home_penalties}-{away_penalties}"
        )


def parse_prediction(
    raw_value: Any,
    match_name: str,
    row: int,
    phase: str,
    warnings: list[str],
    adjacent_penalties: tuple[int, int] | None = None,
) -> dict[str, Any] | None:
    raw = text(raw_value)
    if not raw:
        return None

    match = SCORE_RE.search(raw)
    if not match:
        warnings.append(f"Fila {row}: predicción no interpretable: {raw!r}")
        return {
            "source_row": row,
            "match": match_name or None,
            "prediction_raw": raw,
            "parse_status": "unparsed",
        }

    home_team, away_team = split_match(match_name)
    home_score = int(match.group("home"))
    away_score = int(match.group("away"))
    outcome = match.group("outcome").upper()
    expected_outcome = "1" if home_score > away_score else "2" if away_score > home_score else "X"

    item: dict[str, Any] = {
        "source_row": row,
        "match": match_name or None,
        "home_team": home_team,
        "away_team": away_team,
        "outcome": outcome,
        "home_score": home_score,
        "away_score": away_score,
        "prediction_raw": raw,
        "parse_status": "parsed",
    }

    if not home_team or not away_team:
        warnings.append(f"Fila {row}: no se pudo separar el partido {match_name!r}")

    if outcome != expected_outcome:
        warnings.append(
            f"Fila {row}: signo {outcome!r} no coincide con el marcador "
            f"{home_score}-{away_score}"
        )

    tail = match.group("tail").strip()
    parsed_penalties: tuple[int, int] | None = None
    if tail:
        for penalty_re in PENALTY_RES:
            penalty_match = penalty_re.match(tail)
            if penalty_match:
                parsed_penalties = (
                    int(penalty_match.group("home")),
                    int(penalty_match.group("away")),
                )
                break
        else:
            item["parse_status"] = "parsed_with_unparsed_suffix"
            warnings.append(f"Fila {row}: sufijo/penales no interpretables: {tail!r}")

    if phase != "group_stage" and outcome == "X":
        penalties = parsed_penalties or adjacent_penalties
        if penalties is not None:
            add_penalties(item, penalties[0], penalties[1], row, warnings)
        elif not tail:
            warnings.append(f"Fila {row}: empate eliminatorio sin penales informados")
    elif parsed_penalties is not None:
        add_penalties(item, parsed_penalties[0], parsed_penalties[1], row, warnings)
    elif phase != "group_stage" and home_score == away_score:
        warnings.append(f"Fila {row}: empate eliminatorio sin penales informados")

    return item


def phase_from_header(label: str) -> str | None:
    for header, phase in PHASE_HEADERS.items():
        if header in label:
            return phase
    return None


def mapped_key(label: str, mapping: dict[str, str]) -> str | None:
    for prefix, key in mapping.items():
        if label.startswith(prefix):
            return key
    return None


def convert_excel(
    excel_path: Path,
    player_name: str | None = None,
    output_dir: Path | None = None,
    public_id: str | None = None,
    public_name: str | None = None,
    source_file_name: str | None = None,
) -> ConversionResult:
    excel_path = excel_path.expanduser().resolve()
    if not excel_path.is_file():
        raise FileNotFoundError(f"No existe el archivo Excel: {excel_path}")

    try:
        workbook = load_workbook(excel_path, data_only=True, read_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise RuntimeError(f"No se pudo abrir el Excel: {exc}") from exc

    if SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f'El Excel no contiene la hoja requerida "{SHEET_NAME}"')

    sheet = workbook[SHEET_NAME]
    penalty_lookup = build_penalty_lookup(workbook)
    detected_player = text(sheet["C5"].value)
    final_player_name = text(player_name) or detected_player
    if not final_player_name:
        raise RuntimeError("No se pudo detectar el jugador; usa --player")

    warnings: list[str] = []
    if player_name and detected_player and text(player_name) != detected_player:
        warnings.append(
            f"El nombre indicado ({text(player_name)!r}) reemplaza el de C5 "
            f"({detected_player!r})"
        )

    predictions: dict[str, list[dict[str, Any]]] = {phase: [] for phase in PHASES}
    honor_roll: dict[str, str | None] = {key: None for key in HONOR_LABELS.values()}
    awards: dict[str, str | None] = {key: None for key in AWARD_LABELS.values()}
    qualified_teams: dict[str, list[str]] = {
        phase: [] for phase in QUALIFIED_LABELS.values()
    }
    group_positions: dict[str, list[str]] = {}
    current_phase: str | None = None

    for row in range(1, sheet.max_row + 1):
        identifier = text(sheet.cell(row, 1).value)
        description = text(sheet.cell(row, 2).value)
        value = text(sheet.cell(row, 3).value)
        label = normalized_label(description)

        new_phase = phase_from_header(label)
        if new_phase:
            current_phase = new_phase
            continue

        honor_key = mapped_key(label, HONOR_LABELS)
        if honor_key:
            honor_roll[honor_key] = value or None
            continue

        award_key = mapped_key(label, AWARD_LABELS)
        if award_key:
            awards[award_key] = value or None
            continue

        group_match = re.match(r"^([A-L])([1-3])$", identifier)
        if group_match and value:
            parsed = parse_prediction(value, description, row, "group_stage", warnings)
            if parsed:
                parsed["identifier"] = identifier
                parsed["group"] = group_match.group(1)
                parsed["matchday"] = int(group_match.group(2))
                parsed["match_key"] = (
                    f"group_stage|{parsed.get('home_team')}|{parsed.get('away_team')}"
                )
                predictions["group_stage"].append(parsed)
            continue

        position_match = re.match(r"^([1-4])O GRUPO ([A-L])$", label)
        if position_match and value:
            group_positions.setdefault(position_match.group(2), []).append(value)
            continue

        qualified_key = mapped_key(label, QUALIFIED_LABELS)
        if qualified_key and value:
            qualified_teams[qualified_key].append(value)
            continue

        if value and current_phase != "group_stage" and (
            SCORE_RE.search(value) or "-" in description
        ):
            if not current_phase or current_phase == "group_stage":
                warnings.append(f"Fila {row}: partido eliminatorio sin fase reconocida")
                continue
            score_match = SCORE_RE.search(value)
            adjacent_penalties = None
            if score_match:
                home_team, away_team = split_match(description)
                if home_team and away_team:
                    adjacent_penalties = penalty_lookup.get(
                        (
                            home_team,
                            away_team,
                            int(score_match.group("home")),
                            int(score_match.group("away")),
                        )
                    )
            parsed = parse_prediction(
                value,
                description,
                row,
                current_phase,
                warnings,
                adjacent_penalties,
            )
            if parsed:
                parsed["identifier"] = identifier or None
                parsed["match_key"] = (
                    f"{current_phase}|{parsed.get('home_team')}|{parsed.get('away_team')}"
                )
                predictions[current_phase].append(parsed)

    workbook.close()

    for group in "ABCDEFGHIJKL":
        teams = group_positions.get(group, [])
        if len(teams) != 4:
            warnings.append(f"Grupo {group}: se esperaban 4 posiciones y se encontraron {len(teams)}")

    for phase, expected in EXPECTED_MATCH_COUNTS.items():
        found = len(predictions[phase])
        if found != expected:
            warnings.append(
                f"Fase {phase}: se esperaban {expected} partidos y se encontraron {found}"
            )

    for phase, expected in EXPECTED_QUALIFIED_COUNTS.items():
        found = len(qualified_teams[phase])
        if found != expected:
            warnings.append(
                f"Clasificados {phase}: se esperaban {expected} equipos y se encontraron {found}"
            )

    for key, value in honor_roll.items():
        if value is None:
            warnings.append(f"Cuadro de honor ausente o vacío: {key}")

    for key, value in awards.items():
        if value is None:
            warnings.append(f"Premio ausente o vacío: {key}")

    is_public_build = bool(text(public_id) or text(public_name))
    player_id = text(public_id) or (
        slugify(text(public_name)) if is_public_build else slugify(final_player_name)
    )
    output_player_name = text(public_name) or (
        player_id.upper() if is_public_build else final_player_name
    )
    output_dir = (output_dir or PROJECT_ROOT / "data" / "players").resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{player_id}.json"

    player_payload = {"id": player_id, "name": output_player_name}
    if is_public_build:
        player_payload["alias"] = output_player_name

    payload = {
        "schema_version": 1,
        "player": player_payload,
        "source": {
            "file": source_file_name
            or (f"player-source-{player_id}.xlsx" if is_public_build else excel_path.name),
            "sheet": SHEET_NAME,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        },
        "predictions": predictions,
        "group_positions": group_positions,
        "qualified_teams": qualified_teams,
        "honor_roll": honor_roll,
        "awards": awards,
        "warnings": warnings,
    }
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    return ConversionResult(
        player_id=player_id,
        player_name=output_player_name,
        output_path=output_path,
        match_count=sum(len(items) for items in predictions.values()),
        award_count=sum(value is not None for value in awards.values()),
        warnings=warnings,
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Convierte pronósticos de la hoja Pool a JSON normalizado."
    )
    parser.add_argument("excel", type=Path, help="Ruta al archivo .xlsx")
    parser.add_argument("--player", help="Nombre del jugador (opcional; por defecto usa C5)")
    parser.add_argument("--public-id", help="Identificador público para el JSON, por ejemplo dz-01")
    parser.add_argument("--public-name", help="Alias público visible, por ejemplo DZ-01")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=PROJECT_ROOT / "data" / "players",
        help="Carpeta de salida (por defecto: data/players)",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        result = convert_excel(
            args.excel,
            args.player,
            args.output_dir,
            args.public_id,
            args.public_name,
        )
    except (FileNotFoundError, RuntimeError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    for warning in result.warnings:
        print(f"WARNING: {warning}", file=sys.stderr)

    print("\nConversión completada")
    print(f"Jugador: {result.player_name} ({result.player_id})")
    print(f"Partidos: {result.match_count}")
    print(f"Premios: {result.award_count}")
    print(f"Warnings: {len(result.warnings)}")
    print(f"Salida: {result.output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
